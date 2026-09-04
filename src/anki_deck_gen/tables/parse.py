"""Разбор трёх Источников в одну Таблицу: xlsx, csv, вставленный текст.

Все три сходятся в `_make_sheet`: заголовок уже приведён к каноническим именам,
дальше строки одинаковы. У xlsx лист = вкладка, у csv и текста — один безымянный.
Номера строк — как их видит человек в редакторе (заголовок — строка 1), чтобы
«строка 37» в сообщении бота совпадала с тем, что он откроет в таблице.
"""

import csv
import io
import re
from collections.abc import Iterable
from typing import Any

from openpyxl import load_workbook

from anki_deck_gen.domain import COL_A, COL_DECK, COL_Q, COL_TAGS, Row, Sheet, Table
from anki_deck_gen.errors import TableUnreadable
from anki_deck_gen.tables.headers import normalize_header, require_qa

# Пометка строки вставленного текста без разделителя. Живёт в `extra`, а не в
# отдельном поле Row: это особенность одного Источника, а не словаря домена.
# validate.py превращает её в Problem.NO_SEPARATOR, apply() снимает при правке.
NO_SEPARATOR_MARK = "_no_separator"

# Прямой слэш с пробелами или табуляция. Тире было плохим разделителем: это
# типографский знак, он живёт внутри самих фраз («Москва — столица»), и такая
# строка резалась посередине. Пробелы вокруг слэша обязательны — иначе пострадали
# бы «и/или» и «км/ч».
TEXT_SEPARATOR = re.compile(r"\s+/\s+|\t")

_CSV_DELIMITERS = ",;\t"
_CSV_SNIFF_SAMPLE = 4096


def parse_xlsx(data: bytes, *, title: str | None = None) -> Table:
    """Книга Excel / экспорт Google Sheets. Каждая видимая непустая вкладка — Лист."""
    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl бросает и zipfile.BadZipFile, и KeyError, и свои
        raise TableUnreadable(
            f"Файл не открывается как таблица Excel ({type(exc).__name__})."
        ) from exc

    sheets: list[Sheet] = []
    try:
        for worksheet in workbook.worksheets:
            # Скрытые вкладки — черновики и расчёты; человек их не видит и колоды
            # из них не ждёт.
            if worksheet.sheet_state != "visible":
                continue
            rows_iter = worksheet.iter_rows(values_only=True)
            header_cells = next(rows_iter, None)
            if header_cells is None:
                continue  # вкладка без единой строки
            header = [_cell(value) for value in header_cells]
            if not any(header):
                continue  # первая строка пустая — вкладка считается пустой
            canonical = [normalize_header(cell) if cell else "" for cell in header]
            require_qa(
                frozenset(c for c in canonical if c), sheet=worksheet.title, first_row=header
            )
            data_rows = (
                (number, [_cell(value) for value in cells])
                for number, cells in enumerate(rows_iter, start=2)
            )
            sheets.append(_make_sheet(canonical, data_rows, sheet=worksheet.title))
    finally:
        workbook.close()

    if not sheets:
        raise TableUnreadable("В файле нет ни одной видимой вкладки с данными.")
    return Table(sheets=tuple(sheets), title=title)


def parse_csv(data: bytes, *, title: str | None = None) -> Table:
    """CSV из Excel или Google Sheets: BOM, `;` в русской локали, кавычки — всё штатно."""
    try:
        # utf-8-sig: съедает BOM, который Excel ставит в начало UTF-8 файла.
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise TableUnreadable(
            "Файл не в кодировке UTF-8. Сохраните таблицу как «CSV UTF-8» или пришлите .xlsx."
        ) from exc

    try:
        delimiter = (
            csv.Sniffer().sniff(text[:_CSV_SNIFF_SAMPLE], delimiters=_CSV_DELIMITERS).delimiter
        )
    except csv.Error:
        delimiter = ","

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    header = next(reader, None)
    if header is None or not any(cell.strip() for cell in header):
        raise TableUnreadable("Файл пустой: нет строки заголовка.")
    header = [cell.strip() for cell in header]
    canonical = [normalize_header(cell) if cell else "" for cell in header]
    require_qa(frozenset(c for c in canonical if c), sheet=None, first_row=header)
    # Номер записи, а не физической строки файла: поле в кавычках может тянуться
    # на несколько строк, и тогда нумерация «по переводам строки» разошлась бы с
    # тем, что показывает табличный редактор.
    data_rows = (
        (number, [cell.strip() for cell in cells]) for number, cells in enumerate(reader, start=2)
    )
    return Table(sheets=(_make_sheet(canonical, data_rows, sheet=None),), title=title)


def parse_text(text: str) -> Table:
    """Вставленный текст: строка = запись, `вопрос / ответ`. Заголовка нет.

    Строка без разделителя не отбрасывается, а помечается: это Проблемная строка,
    её покажут в диалоге правки с номером — тем же, что в окне ввода.
    """
    rows: list[Row] = []
    for number, line in enumerate(text.splitlines(), start=1):
        if not line.strip():
            continue
        parts = TEXT_SEPARATOR.split(line, maxsplit=1)
        if len(parts) == 2:
            rows.append(
                Row(number=number, sheet=None, question=parts[0].strip(), answer=parts[1].strip())
            )
        else:
            rows.append(
                Row(
                    number=number,
                    sheet=None,
                    question=line.strip(),
                    answer="",
                    extra={NO_SEPARATOR_MARK: "1"},
                )
            )
    sheet = Sheet(name=None, columns=frozenset({COL_Q, COL_A}), rows=tuple(rows))
    return Table(sheets=(sheet,), title=None)


def looks_like_text_table(text: str) -> bool:
    """Похоже ли сообщение на Таблицу текстом.

    Хотя бы одна строка с разделителем, и таких — не меньше половины. Двух строк
    больше не требуем: пока разделителем было тире, одна фраза с тире посреди
    слишком легко сходила за колоду; ` / ` с пробелами в обычной фразе почти не
    встречается, и одна строка — законная колода из одной записи. Прозаический
    абзац с одним слэшем правило по-прежнему отсекает: половины не набирается.
    А строка без разделителя среди нескольких с ним — Проблемная строка, которую
    человек поправит в диалоге, а не повод молчать.
    """
    lines = [line for line in text.splitlines() if line.strip()]
    if not lines:
        return False
    with_separator = sum(1 for line in lines if TEXT_SEPARATOR.search(line))
    return with_separator >= 1 and with_separator * 2 >= len(lines)


def _cell(value: Any) -> str:
    """Ячейка → строка. Числа без дробной части — без «.0», иначе «5.0 мл» вместо «5 мл»."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _make_sheet(
    canonical: list[str],
    data_rows: Iterable[tuple[int, list[str]]],
    *,
    sheet: str | None,
) -> Sheet:
    columns = frozenset(name for name in canonical if name)
    rows: list[Row] = []
    for number, cells in data_rows:
        if not any(cells):
            continue  # пустые строки-разделители в таблицах — обычное дело
        values: dict[str, str] = {}
        for name, value in zip(canonical, cells, strict=False):
            if name:
                values[name] = value
        tags = tuple(tag.strip() for tag in values.get(COL_TAGS, "").split(",") if tag.strip())
        extra = {
            name: value
            for name, value in values.items()
            if name not in (COL_Q, COL_A, COL_DECK, COL_TAGS)
        }
        rows.append(
            Row(
                number=number,
                sheet=sheet,
                question=values.get(COL_Q, ""),
                answer=values.get(COL_A, ""),
                deck=values.get(COL_DECK) or None,
                tags=tags,
                extra=extra,
            )
        )
    return Sheet(name=sheet, columns=columns, rows=tuple(rows))
